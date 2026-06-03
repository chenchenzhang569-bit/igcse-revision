#!/usr/bin/env python3
"""生成 IGMaster 配置报告 Excel"""

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import datetime

wb = Workbook()

# 样式
header_font = Font(bold=True, color="FFFFFF", size=11)
header_fill = PatternFill(start_color="001C71", end_color="001C71", fill_type="solid")
section_fill = PatternFill(start_color="E8EAF6", end_color="E8EAF6", fill_type="solid")
section_font = Font(bold=True, size=12, color="001C71")
thin_border = Border(
    left=Side(style='thin'), right=Side(style='thin'),
    top=Side(style='thin'), bottom=Side(style='thin')
)

def add_sheet(wb, title, headers, rows, col_widths=None):
    ws = wb.create_sheet(title=title)
    # Header
    for i, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=i, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center')
        cell.border = thin_border
    # Rows
    for r, row in enumerate(rows, 2):
        for c, val in enumerate(row, 1):
            cell = ws.cell(row=r, column=c, value=val)
            cell.border = thin_border
            cell.alignment = Alignment(wrap_text=True, vertical='top')
    # Column widths
    if col_widths:
        for i, w in enumerate(col_widths, 1):
            ws.column_dimensions[get_column_letter(i)].width = w
    # Freeze top row
    ws.freeze_panes = 'A2'
    return ws

# ══════════════════════════════════
# Sheet 1: 项目概览
# ══════════════════════════════════
ws = wb.active
ws.title = "项目概览"
ws['A1'] = 'IGMaster 完整配置报告'
ws['A1'].font = Font(bold=True, size=16, color="001C71")
ws['A2'] = f'生成日期: {datetime.date.today().strftime("%Y-%m-%d")}'
ws['A2'].font = Font(size=10, color="666666")

add_sheet(wb, "部署架构", ['平台', '节点', '用途', '分支', '费用'],
    [
        ['Vercel', '全球 CDN', '开发/测试', 'main', 'Hobby (免费)'],
        ['Zeabur', '香港', '正式上线', 'production', 'Developer ($5/月)'],
        ['Supabase', '美国', '数据库+Storage+Auth', '—', 'Pro ($25/月)'],
        ['阿里云', '中国大陆', '域名 DNS', '—', '按量付费'],
    ], [15, 15, 25, 15, 20]
)

# ══════════════════════════════════
# Sheet 3: 环境变量
# ══════════════════════════════════
add_sheet(wb, "环境变量", ['变量名', '必填', '说明', '备注'],
    [
        ['NEXT_PUBLIC_SUPABASE_URL', '是', 'Supabase 项目 URL', '已在 Vercel/Zeabur 配置'],
        ['NEXT_PUBLIC_SUPABASE_ANON_KEY', '是', '公开 anon key', '已在 Vercel/Zeabur 配置'],
        ['SUPABASE_SERVICE_ROLE_KEY', '是', '管理密钥（勿泄露）', '已在本机脚本使用'],
        ['NEXT_PUBLIC_SITE_URL', '是', '网站地址', 'https://igmaster.org'],
        ['SME_EMAIL', '有默认值', 'SME 登录邮箱', 'inspiringchermann@vmail.dev'],
        ['SME_PASSWORD', '有默认值', 'SME 登录密码', '已有默认值可工作'],
        ['ALIPAY_APP_ID', '（搁置）', '支付宝 APPID', '等注册个体户'],
        ['ALIPAY_PRIVATE_KEY', '（搁置）', '支付宝私钥', '等注册个体户'],
        ['ALIPAY_PUBLIC_KEY', '（搁置）', '支付宝公钥', '等注册个体户'],
        ['ALIPAY_GATEWAY', '（搁置）', '支付宝网关', '默认为沙箱'],
        ['ALIPAY_NOTIFY_URL', '（搁置）', '异步通知 URL', '等注册个体户'],
        ['ALIPAY_RETURN_URL', '（搁置）', '同步跳转 URL', '等注册个体户'],
    ], [30, 15, 30, 30]
)

# ══════════════════════════════════
# Sheet 4: 备份方案
# ══════════════════════════════════
add_sheet(wb, "备份方案", ['备份类型', '频率', '时间', '保存周期', '脚本命令', '输出目录'],
    [
        ['数据库', '每日', '03:00', '30天', 'backup_all.py --type db', '~/backups/db/'],
        ['Storage 文件', '每周', '周日 04:00', '2份', 'backup_all.py --type storage', '~/backups/storage/'],
        ['代码', '每周', '周日 04:00', '7天', 'backup_all.py --type code', '~/backups/code/'],
        ['完整备份', '每周', '周日 04:00', '同各子项', 'backup_all.py', '~/backups/'],
    ], [15, 10, 12, 10, 30, 20]
)

# 数据库表大小
add_sheet(wb, "数据库表", ['表名', '数据行数', '说明'],
    [
        ['exam_boards', 2, '考试局'],
        ['subjects', 11, '科目'],
        ['topics', 143, '主题'],
        ['subtopics', 390, '子主题'],
        ['questions', 5379, '题目（最大表）'],
        ['notes', 450, '笔记'],
        ['past_papers', 4137, '历年试卷'],
        ['purchases', 21, '购买记录'],
        ['profiles', 17, '用户资料'],
        ['error_reports', 10, '错误报告'],
        ['login_events', 13, '登录事件'],
        ['user_roles', 1, '用户角色'],
        ['app_config', 3, '应用配置（含 Resend Key）'],
        ['mock_exam_sets', 49, '模拟考套题'],
        ['mock_exam_papers', 97, '模拟考试卷'],
        ['mock_exam_questions', 1699, '模拟考题目'],
        ['auth_users', '17 人', '认证用户'],
    ], [20, 12, 25]
)

# ══════════════════════════════════
# Sheet 5: 安全审计
# ══════════════════════════════════
add_sheet(wb, "安全审计", ['项目', '状态', '说明'],
    [
        ['代码硬编码密钥', '✅ 无', '所有密钥通过环境变量读取'],
        ['Service role key 使用', '✅ 仅服务端', '只在 API routes 使用'],
        ['Admin 路由验证', '✅ 已修复', 'errors 和 download-sme-ms 已加验证'],
        ['SME 密码硬编码', '✅ 已修复', '改为环境变量读取，有 fallback 值'],
        ['付费内容保护', '✅', 'access.ts 服务端验证购买'],
        ['中间件路由保护', '✅', '未登录重定向到 /login'],
        ['purchases RLS', '✅ 正确', '只能看自己的购买'],
        ['公开表 RLS', '✅ 正确', '所有人都能读'],
        ['Storage 权限', '✅', '公开桶可匿名读，私有桶受保护'],
    ], [25, 12, 40]
)

# ══════════════════════════════════
# Sheet 6: Storage 桶
# ══════════════════════════════════
add_sheet(wb, "Storage桶", ['桶名', '公开', '用途'],
    [
        ['past-papers', '是', '历年真题 PDF'],
        ['mock-exams', '是', '模拟考试卷'],
        ['notes-pdfs', '是', '笔记 PDF'],
        ['sme-raw-backup', '否', 'SME 原始数据备份'],
        ['sme-images', '是', 'SME 截图图片'],
        ['scripts', '是', '辅助脚本'],
    ], [20, 8, 30]
)

# ══════════════════════════════════
# Sheet 7: 待办事项
# ══════════════════════════════════
add_sheet(wb, "待办事项", ['优先级', '事项', '说明'],
    [
        ['高', '合并 main → production', 'auth 修复已推 main，Zeabur 还未部署'],
        ['中', '统一 admin 验证体系', 'requireAdmin vs checkAdmin 两套'],
        ['中', '修复 error_reports RLS', '递归问题，但不影响 admin 功能'],
        ['低', '支付宝支付上线', '需注册个体户'],
        ['低', 'SMTP SPF 配置', '需阿里云 AccessKey'],
        ['低', '域名绑定 Vercel', 'Zeabur 宕机时可切换'],
    ], [10, 25, 40]
)

# 保存
path = '/home/ubuntu/igcse-revision/IGMaster_配置报告.xlsx'
wb.save(path)
print(f"✅ Excel 文档生成成功: {path}")
